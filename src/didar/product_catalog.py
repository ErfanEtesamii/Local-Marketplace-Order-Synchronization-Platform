"""
Marketplace product-title -> Didar catalog Code lookup, sourced from a
client-maintained Excel export of the existing Didar product catalog.

WHY THIS EXISTS
----------------
Client feedback (2026-08-29): most products already exist in Didar under
short, internal names ("راستین 1") that have no relationship to the
long, marketing-styled titles marketplaces send (e.g. Digikala's "ست
هدیه مسی فراز هنر مدل راستین کد 1 | چند رنگ | گارانتی اصالت و سلامت
فیزیکی کالا"). Before this module, upsert_product() (product_client.py)
used the marketplace SKU - or, failing that, the raw marketplace title -
as the Didar Code. For these products that means /product/search never
finds the real existing entry, and a brand-new, wrongly-named duplicate
product gets created in Didar on every order.

The client keeps a full export of the Didar catalog in Excel (columns
include عنوان محصول / کد محصول - product title / product Code, see
_TITLE_HEADER/_CODE_HEADER below for the exact header text expected).
This module matches a marketplace item's title against that catalog to
recover the real Code, so the caller (deal_client.py) can search Didar
for the product that's ACTUALLY there instead of guessing at a Code
from the marketplace's own SKU/title.

MATCHING ALGORITHM (client's explicit choice: automatic word-overlap,
not a hand-authored keyword table like category_mapping.py)
-----------------------------------------------------------
A marketplace title is long, marketing-styled text that is a SUPERSET of
the real (short) catalog name's words, not an exact match - e.g. the
example above contains "راستین" and "1" among many other words. Naive
"most words in common wins" scoring gets this wrong: a longer catalog
entry that happens to share MORE words in total ("پک هدیه راستین 1" -
4 words, 3 of which - هدیه/راستین/1 - appear in the marketplace title)
would outscore the actually-correct, shorter entry ("راستین 1" - both
of its 2 words appear), since raw intersection-count rewards longer
candidates rather than correct ones.

Fix: score by CONTAINMENT, not raw overlap - keep only catalog entries
whose entire (normalized) word set is a SUBSET of the marketplace
title's word set, then, among those, pick the one with the MOST words
(the most specific full match). "پک هدیه راستین 1" fails outright here
("پک" never appears in the marketplace title - Digikala's title says
"ست", not "پک") so it's excluded before any scoring happens; "راستین 1"
passes (both its words are present) and wins as the only/best match.

Deliberately conservative: if NOTHING in the catalog is a full word-set
match, this returns None rather than guessing at a partial match - a
wrong Code silently links an order to the wrong product in Didar, which
is worse than falling back to the old SKU/title behaviour (see
deal_client.py's caller for that fallback).

Reuses category_mapping._normalize_fa for the same Persian text quirks
(Arabic ي/ك vs Persian ی/ک, ZWNJ) so a title spelled either way still
matches, instead of duplicating that logic here.

NOT YET CONFIRMED: the client's export may occasionally contain more
than one row whose full word set is tied for "most specific" against a
given marketplace title (e.g. two differently-coded catalog rows with
the exact same title text - seen to happen for manually-entered
duplicates in other Didar data per contact_client.py's incident notes).
When that happens this logs a warning and deterministically picks the
first tied entry in catalog order rather than raising - a strict block
on every tie would stop sync for well-behaved cases where the tie is
merely a genuine duplicate row pointing at the same real product.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from src.didar.category_mapping import _normalize_fa
from src.logger import get_logger

log = get_logger(__name__)

# Exact header text from the client's Excel export (see the sheet's own
# header row). Matched by header NAME, not column position, so the
# client re-ordering columns in a future export doesn't break this.
_TITLE_HEADER = "عنوان محصول"
_CODE_HEADER = "کد محصول"

# Words too generic to help identify a specific catalog item - stripped
# from both sides before tokenizing. Kept intentionally short and
# generic-only (same "conservative over broad" philosophy as
# category_mapping.py's keyword lists) - a stopword that's actually
# meaningful for some product would silently make an unrelated title
# look like a match.
_STOPWORDS = frozenset({"عدد"})

_SEPARATOR_CHARS = "|/\\,،-–—_()[]{}×"
_SEPARATOR_TABLE = str.maketrans({ch: " " for ch in _SEPARATOR_CHARS})

# Persian (۰-۹, U+06F0-06F9) and Arabic-Indic (٠-٩, U+0660-0669) digits ->
# ASCII. Confirmed necessary against the client's real catalog: it mixes
# digit systems within the same sheet (e.g. "قاب خاتم ۱۰×۱۵ ..." uses
# Persian digits, most other rows use ASCII) while marketplace titles
# from the platforms this project talks to consistently use ASCII
# digits (see the "کد 1" example in this module's docstring). Since the
# whole point of this matcher is distinguishing variants BY NUMBER
# ("راستین 1" vs "راستین 16"), a digit-system mismatch would silently
# fail a match that should succeed - normalizing both sides to ASCII
# digits before tokenizing avoids that.
_DIGIT_TABLE = str.maketrans(
    "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
    "01234567890123456789",
)

# BUGFIX (client feedback, 2026-09 - "product names don't match our real
# catalog"): the client's real catalog has many rows where a model name
# is glued directly to its variant number with no space at all, e.g.
# "چاپا4" (Code=4) - while the marketplace title for the exact same
# product writes it as two separate words, e.g. "... مدل چاپا کد 04 ...".
# Confirmed against the real catalog export: ~128 of its ~3,300 rows have
# a Persian letter immediately followed by a digit this way. Without
# splitting these, "چاپا4" tokenizes as ONE token that can never appear
# in a marketplace title (which always has "چاپا" and the number as
# separate words), so containment matching fails outright for an
# otherwise-correct, already-existing catalog product - and the sync
# creates a wrong duplicate product under the raw marketplace title
# instead (see deal_client.py's fallback).
#
# Scoped to PERSIAN letters only (Unicode block \u0600-\u06FF, after
# _normalize_fa has already unified Arabic/Persian variants) - NOT plain
# ASCII letters - so this doesn't touch the deliberate "10x15" vs "10×15"
# distinction already covered by test_persian_digits_in_catalog_match_
# ascii_digits_in_title (an ASCII "x" typed for the multiplication sign
# must still NOT match "×" - unrelated to this glued-Persian-word issue).
_PERSIAN_DIGIT_BOUNDARY = re.compile(
    r"(?<=[\u0600-\u06FF])(?=[0-9])|(?<=[0-9])(?=[\u0600-\u06FF])"
)


@dataclass(frozen=True)
class CatalogMatch:
    code: str
    title: str  # the catalog's own (short) title - prefer this over the
                # marketplace's raw title once matched, since it's what
                # is actually already in Didar.


def _split_glued_persian_digits(token: str) -> list[str]:
    """Split a token where a Persian model name is glued directly to a
    digit run with no separator - see _PERSIAN_DIGIT_BOUNDARY above."""
    return [part for part in _PERSIAN_DIGIT_BOUNDARY.sub(" ", token).split() if part]


def _strip_leading_zeros(token: str) -> str:
    """Normalize a purely-numeric token by stripping leading zeros, e.g.
    "04" -> "4" (client's real catalog confirms this is the same product
    - catalog row "چاپا4"/Code=4 vs marketplace title "... چاپا کد 04
    ..."). A leading-zero-only difference must not fail an otherwise
    exact match. Left untouched for non-numeric tokens, and for a token
    that's ALL zeros (kept as "0" rather than stripped to empty)."""
    if token.isdigit() and len(token) > 1:
        return token.lstrip("0") or "0"
    return token


def _tokenize(text: str) -> frozenset[str]:
    normalized = _normalize_fa(text).translate(_DIGIT_TABLE).translate(_SEPARATOR_TABLE)
    tokens: list[str] = []
    for raw in normalized.split():
        for part in _split_glued_persian_digits(raw):
            normalized_part = _strip_leading_zeros(part)
            if normalized_part and normalized_part not in _STOPWORDS:
                tokens.append(normalized_part)
    return frozenset(tokens)


class ProductCatalog:
    """Loads a client-maintained Excel product catalog once and answers
    title -> Code lookups against it. Holds the whole (small: a few
    thousand rows in practice) catalog in memory - not thread-safe, same
    assumption as DidarProductClient's own per-instance category cache."""

    def __init__(self, xlsx_path: str | Path):
        self._entries: list[tuple[frozenset[str], str, str]] = []  # (tokens, title, code)
        self._load(Path(xlsx_path))

    def _load(self, path: Path) -> None:
        if not path.exists():
            raise FileNotFoundError(
                f"didar: product catalog Excel not found at {path} - set "
                f"DIDAR_PRODUCT_CATALOG_XLSX in .env to the real path, or "
                f"leave it blank to disable catalog-based Code lookup "
                f"entirely (falls back to marketplace SKU/title - see "
                f"deal_client.py)."
            )
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            header = ()
        try:
            title_idx = header.index(_TITLE_HEADER)
            code_idx = header.index(_CODE_HEADER)
        except ValueError as exc:
            raise ValueError(
                f"didar: product catalog Excel at {path} is missing an "
                f"expected column - looked for {_TITLE_HEADER!r} and "
                f"{_CODE_HEADER!r} in the header row, got {list(header)!r}"
            ) from exc

        loaded = 0
        for row in rows:
            if row is None:
                continue
            title = row[title_idx] if title_idx < len(row) else None
            code = row[code_idx] if code_idx < len(row) else None
            if not title or code is None or str(code).strip() == "":
                continue
            code_str = str(code).strip()
            if isinstance(code, float) and code.is_integer():
                # Defensive: the client's real export has Code as text
                # (e.g. "146"), but a future re-export/re-save in Excel
                # could turn a numeric-looking column into actual
                # numbers, silently producing "146.0" instead of "146"
                # (which would never match Didar's own Code string).
                code_str = str(int(code))
            tokens = _tokenize(str(title))
            if not tokens:
                continue
            self._entries.append((tokens, str(title).strip(), code_str))
            loaded += 1
        log.info("didar: loaded %d product catalog entries from %s", loaded, path)

    def match(self, platform_title: str) -> CatalogMatch | None:
        """Best catalog match for a marketplace item's title, or None if
        nothing in the catalog is fully word-set-contained in it - see
        module docstring for the containment rule and why it's used
        instead of raw word-overlap scoring."""
        if not platform_title:
            return None
        title_tokens = _tokenize(platform_title)
        if not title_tokens:
            return None

        candidates = [
            (tokens, title, code)
            for tokens, title, code in self._entries
            if tokens <= title_tokens  # full subset containment
        ]
        if not candidates:
            return None

        candidates.sort(key=lambda c: len(c[0]), reverse=True)
        best_len = len(candidates[0][0])
        tied = [c for c in candidates if len(c[0]) == best_len]
        if len(tied) > 1:
            log.warning(
                "didar: ambiguous catalog match for title %r - %d entries "
                "tied at %d matched words (%s) - using the first one found "
                "in the catalog",
                platform_title, len(tied), best_len,
                [t for _, t, _ in tied],
            )
        _, title, code = candidates[0]
        return CatalogMatch(code=code, title=title)
