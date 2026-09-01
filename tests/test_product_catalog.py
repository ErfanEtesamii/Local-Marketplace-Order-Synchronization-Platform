from openpyxl import Workbook

from src.didar.product_catalog import ProductCatalog

_HEADER = [
    "_type", "عنوان محصول", "دسته بندی محصول", "کد دیدار محصول", "کد محصول",
]


def _make_catalog(tmp_path, rows):
    """Build a tiny .xlsx fixture with the same header shape as the
    client's real export, containing only the given (title, code) rows,
    and return a ProductCatalog loaded from it."""
    wb = Workbook()
    ws = wb.active
    ws.append(_HEADER)
    for title, code in rows:
        ws.append(["Product", title, None, 0, code])
    path = tmp_path / "catalog.xlsx"
    wb.save(path)
    return ProductCatalog(path)


def test_short_specific_entry_wins_over_longer_superset_entry(tmp_path):
    """Regression test for the real client example: a Digikala title
    like 'ست هدیه مسی فراز هنر مدل راستین کد 1 | چند رنگ | گارانتی
    اصالت و سلامت فیزیکی کالا' must resolve to the short catalog entry
    'راستین 1', NOT the longer 'پک هدیه راستین 1' - even though the
    longer one shares MORE words in total ("هدیه" + "راستین" + "1" vs
    just "راستین" + "1") - because "پک" never appears in the
    marketplace title (which says "ست", not "پک"), so it fails full
    containment and must be excluded outright.
    """
    catalog = _make_catalog(tmp_path, [
        ("راستین 1", "146"),
        ("پک هدیه راستین 1", "28200"),
        ("راستین 16", "161"),
    ])
    match = catalog.match(
        "ست هدیه مسی فراز هنر مدل راستین کد 1 | چند رنگ | "
        "گارانتی اصالت و سلامت فیزیکی کالا"
    )
    assert match is not None
    assert match.code == "146"
    assert match.title == "راستین 1"


def test_trailing_number_prevents_wrong_variant_match(tmp_path):
    """'راستین 16' must not match a title that only mentions '1', not
    '16' - "1" and "16" are different tokens, not a substring hit."""
    catalog = _make_catalog(tmp_path, [
        ("راستین 1", "146"),
        ("راستین 16", "161"),
    ])
    match = catalog.match("ست هدیه راستین کد 1 چند رنگ")
    assert match is not None
    assert match.code == "146"


def test_no_match_returns_none(tmp_path):
    catalog = _make_catalog(tmp_path, [("راستین 1", "146")])
    assert catalog.match("یک محصول کاملا نامرتبط بدون کلیدواژه") is None


def test_empty_title_returns_none(tmp_path):
    catalog = _make_catalog(tmp_path, [("راستین 1", "146")])
    assert catalog.match("") is None
    assert catalog.match(None) is None


def test_arabic_yeh_normalization_still_matches(tmp_path):
    # Arabic yeh (ي) in the catalog title vs Persian yeh (ی) in the
    # marketplace title must still match - same normalization as
    # category_mapping.py.
    catalog = _make_catalog(tmp_path, [("را\u064aعلي 1", "9")])
    match = catalog.match("محصول رایعلی 1 با گارانتی")
    assert match is not None
    assert match.code == "9"


def test_persian_digits_in_catalog_match_ascii_digits_in_title(tmp_path):
    """Regression test: the real client catalog mixes digit systems
    (e.g. 'قاب خاتم \u06f0\u06f9\u064817\u06f5\u064a' style rows written
    with Persian digits) while marketplace titles use ASCII digits. The
    variant number is exactly what distinguishes catalog entries, so a
    digit-system mismatch must not block an otherwise-exact match."""
    catalog = _make_catalog(tmp_path, [("قاب خاتم \u06f1\u06f0\u00d7\u06f1\u06f5", "500")])
    match = catalog.match("قاب خاتم 10x15 نیم توره پلاک نقشه ایران")
    assert match is None  # "x" (ascii letter) is not a recognized separator - by design
    match2 = catalog.match("قاب خاتم 10×15 نیم توره پلاک نقشه ایران")
    assert match2 is not None
    assert match2.code == "500"


def test_float_code_normalized_to_plain_integer_string(tmp_path):
    """Defensive: if a future Excel re-export stores Code as a real
    number (146.0) instead of text ("146"), it must still resolve to
    the plain "146" Didar actually uses as its Code - not "146.0"."""
    wb = Workbook()
    ws = wb.active
    ws.append(_HEADER)
    ws.append(["Product", "راستین 1", None, 0, 146.0])
    path = tmp_path / "catalog.xlsx"
    wb.save(path)
    catalog = ProductCatalog(path)
    match = catalog.match("ست هدیه راستین کد 1 چند رنگ")
    assert match is not None
    assert match.code == "146"


def test_glued_persian_word_and_digit_matches_separated_marketplace_title(tmp_path):
    """Regression test for a real production issue (client feedback,
    2026-09 - "product names don't match our real catalog"): the
    client's actual catalog export has rows like "چاپا4" (model name
    glued directly to its number, no space), while marketplace titles
    for the exact same product always write it as separate words plus
    a leading zero, e.g. "... مدل چاپا کد 04 ...". Confirmed against
    the real catalog export: ~128 of its ~3,300 rows have this glued
    letter+digit pattern, each one silently failing to match and
    causing a wrong duplicate product to be created under the raw
    marketplace title instead (see deal_client.py's fallback)."""
    catalog = _make_catalog(tmp_path, [
        ("چاپا4", "4"),
        ("چاپا1", "367"),
        ("چاپا اعلا 1", "36336300430001"),
    ])
    match = catalog.match(
        "رومیزی قلمکار مدل چاپا کد 04 | چند رنگ | "
        "گارانتی اصالت و سلامت فیزیکی کالا"
    )
    assert match is not None
    assert match.code == "4"


def test_missing_expected_columns_raises(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["_type", "عنوان اشتباه", "کد اشتباه"])
    ws.append(["Product", "چیزی", "1"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    try:
        ProductCatalog(path)
        assert False, "expected ValueError for missing columns"
    except ValueError as exc:
        assert "عنوان محصول" in str(exc)


def test_missing_file_raises_clear_error(tmp_path):
    missing = tmp_path / "does-not-exist.xlsx"
    try:
        ProductCatalog(missing)
        assert False, "expected FileNotFoundError"
    except FileNotFoundError as exc:
        assert "DIDAR_PRODUCT_CATALOG_XLSX" in str(exc)
